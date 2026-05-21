"""S6 — Output template structure preserved (spec §15).

After write_output runs, the 7 output sheets must:
1. Have the SAME sheet names as the template.
2. Have the SAME row-2 headers as the template (banner stays at row 1).
3. Be openable in Excel / LibreOffice without warnings (i.e. valid xlsx).

This test runs Quick on the 1k subset (cheapest solve) and inspects the
resulting output.xlsx.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.methods.quick import QuickMethod
from backend.output_writer import write_output


@pytest.mark.timeout(15 * 60)
def test_output_excel_matches_template(small_problem, template_path, tmp_path):
    result = QuickMethod().solve(small_problem, time_cap_sec=300.0, gap_target=0.03, threads=4)
    assert result.status in ("optimal", "gap_reached", "time_limit"), (
        f"Quick failed on 1k subset: status={result.status}"
    )

    out_file = tmp_path / "test_output.xlsx"
    write_output(result, small_problem, out_file, run_id="TEST_S6",
                 template_path=template_path)

    wb = load_workbook(out_file)
    template = load_workbook(template_path)

    # Sheet names match
    assert wb.sheetnames == template.sheetnames, (
        f"Output sheet names {wb.sheetnames} differ from template {template.sheetnames}"
    )

    # Row-2 headers match for every sheet
    for sh in template.sheetnames:
        for col in range(1, template[sh].max_column + 1):
            tmpl_h = template[sh].cell(2, col).value
            out_h = wb[sh].cell(2, col).value
            if tmpl_h is None and out_h is None:
                continue
            assert tmpl_h == out_h, (
                f"Header mismatch in {sh} col {col}: template={tmpl_h!r}, output={out_h!r}"
            )
