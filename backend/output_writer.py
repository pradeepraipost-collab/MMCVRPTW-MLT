"""Write the 7-sheet output Excel from a solved MethodResult (spec §11).

Strategy: load ``MMCVRPTW_MLT_OutputTemplate_V4.xlsx`` with openpyxl, KEEP row 1
banner and row 2 headers intact, write data starting at row 3. For
Solve_Summary fill labelled cells in place; for Recommendations rewrite the
merged-cell layout with the six cards. Save to runs/<run_id>/output.xlsx.

The template uses ``data_only=False`` so formulae and styles survive; openpyxl
preserves column widths, fonts, fills via its native styling pipeline.

Test S6 verifies that sheet names and row-2 headers match the template after
this writer runs.
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .methods.base import MethodResult
from .problem import Problem
from . import extract as _extract
from . import recommendations as _reco
from . import cost_comparison as _ccmp


TEMPLATE_PATH = "MMCVRPTW_MLT_OutputTemplate_V4.xlsx"


def write_output(
    result: MethodResult,
    problem: Problem,
    output_path: str | Path,
    run_id: str,
    template_path: str | Path = TEMPLATE_PATH,
) -> Path:
    """Write the 7-sheet output Excel to ``output_path`` and return it."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)

    # Make sure assignments / trips are populated for MILP methods
    _extract.extract_strict_assignments(result, problem)
    timeline = _extract.build_order_timeline(result, problem)
    node_util = _extract.build_node_utilization(result, problem)
    kpis = _extract.order_kpis(result, problem)
    recos = _reco.compute_recommendations(result, problem)
    ccmp = _ccmp.compute_cost_comparison(result, problem)

    _write_solve_summary(wb["Solve_Summary"], result, problem, run_id, kpis, ccmp)
    _write_order_assignment(wb["Order_Assignment"], result, problem)
    _write_order_timeline(wb["Order_Timeline"], timeline)
    _write_trip_plan(wb["Trip_Plan"], result)
    _write_node_utilization(wb["Node_Utilization"], node_util)
    _write_recommendations(wb["Recommendations"], recos)
    _write_cost_comparison(wb["Cost_Comparison"], ccmp, result, problem)

    wb.save(output_path)
    return output_path


# ---------- Sheet writers ----------

def _is_lp_bound(result: MethodResult) -> bool:
    return result.method_id == "lp_bound" or result.status == "lower_bound"


def _method_label(result: MethodResult) -> str:
    return {
        "quick":     "Quick — Aggregated cells",
        "balanced":  "Balanced — Per-FC decomposition",
        "strict":    "Strict — Monolithic per-order MILP",
        "heuristic": "Heuristic — Greedy + 2-opt",
        "lp_bound":  "LP Bound — Relaxation diagnostic",
    }.get(result.method_id, result.method_id)


def _write_solve_summary(ws, result: MethodResult, problem: Problem,
                          run_id: str, kpis: dict, ccmp: dict) -> None:
    """Fill ALL labelled cells in Solve_Summary — value (B), percentage (C),
    and description (D) — from the actual MethodResult. Earlier versions wrote
    column B only and left the template's example percentages / description
    strings (6.7%, 91.2%, "10:12 minutes", "385 orders breached") in place,
    which silently leaked stale numbers from the dev fixture into every run.

    Each row is keyed by its column-A label. We compute the percentage and
    description per row from real data; any label not present in this map
    keeps its template value (those are section headers / static text).
    """
    total_cost = (result.fc_fixed_cost_inr
                  + result.carrier_cost_inr
                  + result.sla_penalty_inr)
    n_orders = max(1, kpis["total_orders"])
    breaches = kpis.get("sla_breaches", 0)

    def _pct(part: float, whole: float) -> str | float:
        if not whole:
            return ""
        return round(100.0 * part / whole, 2)

    def _fmt_wall(sec: float) -> str:
        if sec < 60:
            return f"{sec:.1f} seconds"
        m, s = divmod(int(round(sec)), 60)
        return f"{m}:{s:02d} minutes"

    # row_specs: { label : (value, pct_col_C, desc_col_D) }
    # If pct or desc is None, that column is cleared rather than written.
    row_specs: dict[str, tuple] = {
        "Run_ID":               (run_id, None, None),
        "Run_Date":             (_dt.datetime.now().strftime("%Y-%m-%d"), None, None),
        "Active_Wave":          (f"{problem.order_wave.wave_id} "
                                 f"({_fmt_hr(problem.order_wave.start_hr)}–"
                                 f"{_fmt_hr(problem.order_wave.end_hr)} placement)",
                                 None, None),
        "Profile":              ("Default (1% gap, 60 min cap)", None,
                                 "Single profile — method picker handles speed/quality trade"),
        "Solver":               ("HiGHS 1.7.2 via highspy", None, None),
        "Method_Used":          (_method_label(result), None,
                                 "(Quick | Balanced | Strict | Heuristic | LP Bound)"),
        "Status":               (result.status, None,
                                 "(optimal | gap_reached | time_limit | infeasible)"),
        "Achieved_Gap_pct":     (round(result.achieved_gap_pct, 2)
                                 if result.achieved_gap_pct is not None else "—",
                                 None, "Solver gap at termination"),
        "Target_Gap_pct":       (1.0, None, "Profile target"),
        "Best_Objective_INR":   (round(result.best_objective, 0)
                                 if result.best_objective else "—",
                                 None, "Best feasible solution found"),
        "Best_Lower_Bound_INR": (round(result.best_lower_bound, 0)
                                 if result.best_lower_bound else "—",
                                 None, "Proven lower bound (gap = (obj-bound)/obj)"),
        "Wall_Time_sec":        (round(result.wall_time_sec, 1), None,
                                 _fmt_wall(result.wall_time_sec)),
        "Threads_Used":         (result.threads_used, None, None),

        # Cost breakdown: percentage column is computed from this run's total.
        "FC_Fixed_Cost":        (round(result.fc_fixed_cost_inr, 0),
                                 _pct(result.fc_fixed_cost_inr, total_cost),
                                 f"{len(problem.fcs)} FCs × wave fixed cost"),
        "Carrier_Cost":         (round(result.carrier_cost_inr, 0),
                                 _pct(result.carrier_cost_inr, total_cost),
                                 "FTL + PTL + LTL + Courier rates"),
        "SLA_Penalty_Cost":     (round(result.sla_penalty_inr, 0),
                                 _pct(result.sla_penalty_inr, total_cost),
                                 "Soft penalty on breaches"),
        "TOTAL":                (round(total_cost, 0), 100.0, None),

        # Order KPIs — each percentage is order_count / total_orders.
        "Total_Orders":         (kpis["total_orders"], None, None),
        "Orders_via_Courier":   (kpis["orders_via_courier"],
                                 _pct(kpis["orders_via_courier"], n_orders),
                                 "FC_DIRECT shipments"),
        "Orders_via_HubSpoke":  (kpis["orders_via_hub_spoke"],
                                 _pct(kpis["orders_via_hub_spoke"], n_orders),
                                 "FC→SC→DS"),
        "Orders_via_FC_Direct": (kpis["orders_via_fc_direct"],
                                 _pct(kpis["orders_via_fc_direct"], n_orders),
                                 "FC→DS direct (no SC)"),
        "SLA_Met_pct":          (kpis["sla_met_pct"], None,
                                 f"{breaches} orders breached"),
        "SLA_Breaches":         (breaches, _pct(breaches, n_orders), None),
        "Avg_Cost_per_Order":   (kpis.get("avg_cost_per_order",
                                          round(total_cost / n_orders, 2)),
                                 None, None),

        # Cost comparison section (mirrors the Cost_Comparison sheet so the
        # numbers are consistent across sheets).
        "Courier_Only_Baseline_INR": (round(ccmp.get("baseline_cost_inr", 0), 0),
                                      None, "Every order shipped courier @ ₹85/parcel "
                                            "(oversize-scaled for non-eligible)"),
        "Optimizer_Total_INR":  (round(ccmp.get("optimizer_cost_inr", 0), 0), None, None),
        "Savings_INR":          (round(ccmp.get("savings_inr", 0), 0), None, None),
        "Savings_pct":          (round(ccmp.get("savings_pct", 0), 1), None,
                                 f"Optimizer is {ccmp.get('savings_pct', 0):.1f}% "
                                 + ("cheaper" if ccmp.get('savings_pct', 0) >= 0
                                    else "more expensive")
                                 + " than courier-only"),
    }

    # Load type mix rows (R40-R45 area). Each row: A=label, B=#trips, C=#orders, D=% orders.
    lt_orders = kpis.get("load_type_orders", {})
    lt_trips = kpis.get("load_type_trips", {})
    lt_rows = {
        "FTL":     (lt_trips.get("FTL", 0),     lt_orders.get("FTL", 0),
                    _pct(lt_orders.get("FTL", 0), n_orders)),
        "PTL":     (lt_trips.get("PTL", 0),     lt_orders.get("PTL", 0),
                    _pct(lt_orders.get("PTL", 0), n_orders)),
        "LTL":     (lt_trips.get("LTL", 0),     lt_orders.get("LTL", 0),
                    _pct(lt_orders.get("LTL", 0), n_orders)),
        "Courier": (lt_trips.get("Courier", 0), lt_orders.get("Courier", 0),
                    _pct(lt_orders.get("Courier", 0), n_orders)),
        "FC_Direct (PTL/LTL)": (
            sum(1 for t in result.trips if t.get("lane_type") == "FC_DS"),
            kpis["orders_via_fc_direct"],
            _pct(kpis["orders_via_fc_direct"], n_orders),
        ),
    }

    for r in range(1, ws.max_row + 5):
        label = ws.cell(r, 1).value
        if not isinstance(label, str):
            continue
        key = label.strip()
        if key in row_specs:
            val, pct, desc = row_specs[key]
            ws.cell(r, 2, value=val)
            ws.cell(r, 3, value=pct if pct is not None else None)
            if desc is not None:
                ws.cell(r, 4, value=desc)
        elif key in lt_rows:
            # Load type mix uses a different layout: B=#trips, C=#orders, D=%
            n_trips, n_ord, pct = lt_rows[key]
            ws.cell(r, 2, value=n_trips)
            ws.cell(r, 3, value=n_ord)
            ws.cell(r, 4, value=pct)


def _write_order_assignment(ws, result: MethodResult, problem: Problem) -> None:
    _clear_data_rows(ws, start_row=3)
    if _is_lp_bound(result):
        ws.cell(3, 1, value="Not applicable for LP Bound — the LP relaxation has no integer assignment.")
        return
    for i, a in enumerate(result.assignments, start=3):
        ws.cell(i, 1, value=a["order_id"])
        ws.cell(i, 2, value=a["route_type"])
        ws.cell(i, 3, value=a.get("fc", ""))
        ws.cell(i, 4, value=a.get("sc", ""))
        ws.cell(i, 5, value=a.get("ds", ""))
        ws.cell(i, 6, value=a.get("carrier", ""))
        ws.cell(i, 7, value=a.get("vehicle", ""))
        ws.cell(i, 8, value=a.get("load_type", ""))
        ws.cell(i, 9, value=a.get("trip_id", ""))
        ws.cell(i, 10, value=a.get("dispatch_time", ""))
        ws.cell(i, 11, value=a.get("sla_status", ""))


def _write_order_timeline(ws, timeline: list[dict]) -> None:
    _clear_data_rows(ws, start_row=3)
    if not timeline:
        ws.cell(3, 1, value="Not applicable for LP Bound.")
        return
    for i, t in enumerate(timeline, start=3):
        ws.cell(i, 1, value=t["order_id"])
        ws.cell(i, 2, value=t["fc_dispatch"])
        ws.cell(i, 3, value=t["sc_arrival"])
        ws.cell(i, 4, value=t["sc_dispatch"])
        ws.cell(i, 5, value=t["ds_arrival"])
        ws.cell(i, 6, value=t["ds_dispatch"])
        ws.cell(i, 7, value=t["customer_eta"])
        ws.cell(i, 8, value=t["ops_sla_deadline"])
        ws.cell(i, 9, value=t["sla_status"])


def _write_trip_plan(ws, result: MethodResult) -> None:
    _clear_data_rows(ws, start_row=3)
    if _is_lp_bound(result) or not result.trips:
        ws.cell(3, 1, value="Not applicable for LP Bound." if _is_lp_bound(result) else "No trips.")
        return
    for i, t in enumerate(result.trips, start=3):
        ws.cell(i, 1, value=t.get("trip_id", ""))
        ws.cell(i, 2, value=t.get("lane_type", ""))
        ws.cell(i, 3, value=t.get("lane", ""))
        ws.cell(i, 4, value=t.get("carrier", ""))
        ws.cell(i, 5, value=t.get("vehicle", ""))
        ws.cell(i, 6, value=t.get("load_type", ""))
        ws.cell(i, 7, value=t.get("stop_count", 1))
        ws.cell(i, 8, value=t.get("stop_sequence", ""))
        ws.cell(i, 9, value=t.get("parcels", 0))
        ws.cell(i, 10, value=t.get("distance_km", 0))
        ws.cell(i, 11, value=t.get("fill_weight_pct", 0))
        ws.cell(i, 12, value=t.get("fill_vol_pct", 0))


def _write_node_utilization(ws, node_util: list[dict]) -> None:
    _clear_data_rows(ws, start_row=3)
    for i, n in enumerate(node_util, start=3):
        ws.cell(i, 1, value=n["node_id"])
        ws.cell(i, 2, value=n["node_type"])
        ws.cell(i, 3, value=n["load_parcels"])
        ws.cell(i, 4, value=n["capacity_parcels"])
        ws.cell(i, 5, value=n["utilization_pct"])
        ws.cell(i, 6, value=n["bottleneck_flag"])
        ws.cell(i, 7, value=n["notes"])


def _write_recommendations(ws, recos: dict) -> None:
    """Rewrite the 6-card layout. We don't preserve the template's merged
    cells exactly — instead, we clear the data region and write each card
    contiguously starting at row 3."""
    # Find first data row (skip banner + any intro)
    _clear_data_rows(ws, start_row=3, max_clear_row=80)
    r = 3
    cards = [
        ("CARD 1 — UNDER-UTILIZED FCs", recos["card1_underutilized_fcs"],
            ("fc_id", "city", "region", "load_parcels", "capacity_parcels", "utilization_pct", "suggestion")),
        ("CARD 2 — CARRIERS NEAR CONCENTRATION CAP", recos["card2_carriers_near_cap"],
            ("carrier", "orders_assigned", "pct_of_total", "max_concentration_pct", "buffer_pct", "risk_level")),
        ("CARD 3 — FTL CONSOLIDATION OPPORTUNITIES", recos["card3_ftl_consolidation"],
            ("lane", "current_ptl_ltl_cost", "hypothetical_ftl_cost", "savings_inr", "combined_fill_pct", "action")),
        ("CARD 4 — COURIER vs HUB-SPOKE", recos["card4_courier_vs_hubspoke"],
            ("destination_pincode", "city", "orders", "courier_cost_inr", "hubspoke_cost_inr", "optimizer_choice", "savings_per_order")),
        ("CARD 5 — SLA BREACH RISK", recos["card5_sla_risk"],
            ("fc", "ds", "orders", "breaches", "avg_slack_hr")),
        ("CARD 6 — MULTI-STOP OPPORTUNITIES", recos["card6_multistop_opps"],
            ("city", "single_stop_trip_count", "opportunity", "est_savings_per_consolidation_inr")),
    ]
    for title, rows, cols in cards:
        ws.cell(r, 1, value=title); r += 1
        # header row
        for ci, h in enumerate(cols, start=1):
            ws.cell(r, ci, value=h)
        r += 1
        for row in rows:
            for ci, h in enumerate(cols, start=1):
                ws.cell(r, ci, value=row.get(h, ""))
            r += 1
        r += 1  # blank row between cards


def _write_cost_comparison(ws, ccmp: dict, result: MethodResult,
                            problem: Problem) -> None:
    """Rebuild the Cost_Comparison sheet entirely from the current
    MethodResult. The template ships with example values from a Balanced run
    (₹31.2M baseline, 1.87% gap, 612s wall) that are wrong for every other
    method/run. We clear all data rows below the header and write fresh
    rows that match the current solve.
    """
    # Unmerge any merges that overlap the data region so we can write freely
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= 2:  # keep row-1 banner merge intact
            to_unmerge.append(str(rng))
    for r in to_unmerge:
        ws.unmerge_cells(r)

    # Clear everything from row 2 onwards (we'll re-write headers).
    for r in range(2, ws.max_row + 5):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c, value=None)

    n_orders = max(1, len(problem.orders))
    baseline = ccmp.get("baseline_cost_inr", 0.0)
    optimizer = ccmp.get("optimizer_cost_inr", 0.0)
    savings_inr = ccmp.get("savings_inr", 0.0)
    savings_pct = ccmp.get("savings_pct", 0.0)
    method_label = _method_label(result)
    note_optimiser = (f"{method_label}, status={result.status}, "
                      f"wall={result.wall_time_sec:.1f}s"
                      + (f", gap={result.achieved_gap_pct:.2f}%"
                         if result.achieved_gap_pct is not None else ""))

    # SCENARIO header (row 2)
    ws.cell(2, 1, value="SCENARIO")
    ws.cell(2, 2, value="Orders")
    ws.cell(2, 3, value="Total_Cost_INR")
    ws.cell(2, 4, value="Notes")

    # Scenario rows
    ws.cell(3, 1, value="Courier-Only Baseline")
    ws.cell(3, 2, value=n_orders)
    ws.cell(3, 3, value=round(baseline, 0))
    ws.cell(3, 4, value=("Every order shipped via FC_DIRECT courier at ₹85/parcel "
                         "(oversize-scaled for non-eligible)"))

    ws.cell(4, 1, value="Optimizer Solution")
    ws.cell(4, 2, value=n_orders)
    ws.cell(4, 3, value=round(optimizer, 0))
    ws.cell(4, 4, value=note_optimiser)

    ws.cell(5, 1, value="Savings (INR)")
    ws.cell(5, 3, value=round(savings_inr, 0))

    ws.cell(6, 1, value="Savings (%)")
    ws.cell(6, 3, value=round(savings_pct, 1))
    ws.cell(6, 4, value=(f"Optimizer is {savings_pct:.1f}% "
                         + ("cheaper" if savings_pct >= 0 else "more expensive")
                         + " than courier-only"))

    # BREAKDOWN BY ROUTE TYPE — rebuild from actual breakdown payload.
    ws.cell(8, 1, value="BREAKDOWN BY ROUTE TYPE")
    ws.cell(8, 2, value="Orders")
    ws.cell(8, 3, value="Cost_INR")
    ws.cell(8, 4, value="Avg_INR_per_Order")
    r = 9
    for item in ccmp.get("breakdown_by_route_type", []):
        ws.cell(r, 1, value=f"Optimizer: {item['route_type']}")
        ws.cell(r, 2, value=(item["orders"] if item["orders"] is not None else "—"))
        ws.cell(r, 3, value=round(item["cost_inr"], 0))
        ws.cell(r, 4, value=round(item["avg_inr_per_order"], 1))
        r += 1

    # TAKEAWAYS — also recomputed from current ccmp payload, not stale.
    r += 1
    ws.cell(r, 1, value="TAKEAWAYS")
    r += 1
    for line in ccmp.get("takeaways", []):
        ws.cell(r, 1, value=line)
        r += 1


def _clear_data_rows(ws, start_row: int, max_clear_row: int = 12000) -> None:
    """Clear rows from ``start_row`` down to wherever existing data ends.

    Recommendations sheet has merged cells inside the data region (one per card
    title row). openpyxl raises AttributeError when writing to a non-top-left
    cell of a merge — so we unmerge any range that intersects the data region
    before clearing. Banner merges (row 1) are above ``start_row`` and untouched.
    """
    # Unmerge any range that overlaps the clear region; openpyxl writes will
    # then succeed cell-by-cell.
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= start_row:
            to_unmerge.append(str(rng))
    for r in to_unmerge:
        ws.unmerge_cells(r)

    last = min(ws.max_row, max_clear_row)
    for r in range(start_row, last + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c, value=None)


def _fmt_hr(hr: float) -> str:
    h = int(hr) % 24; m = int(round((hr - int(hr)) * 60))
    return f"{h:02d}:{m:02d}"
