#!/usr/bin/env python3
"""
revise_master_round2.py
========================
Round 2 of master Excel recalibration.

Changes from Round 1:
  1. SLA_Config — distance-aware. Long-haul Metro-to-Metro gets realistic
     deadlines (48h Prime / 72h Standard) instead of the impossible 28h/36h
     that caused 3,725 of 10,000 orders to have negative slack on direct routes.
     Same-City and Same-Region bands unchanged.
  2. Carrier_Master FTL rates — reduced 25-30% from Round 1's overly-conservative
     values. Solver was barely using trucks (carrier_cost only 8% of total)
     because trucks were priced too high vs courier+breach.
  3. Penalty_Config — bumped to moderate-high values (₹5k-15k) so solver
     strongly prefers SLA-meeting routes but doesn't cause infeasibility
     for the genuinely-impossible orders.
  4. Order_Data SLA_hrs / SLA_Tier columns — RECOMPUTED from the new SLA_Config
     to match. Each order's deadline is re-derived based on its FC region and
     DS city; SLA_Tier label is updated to reflect the new band.

Backup: Saves .bak2 copy alongside revised file.

Usage:
    cd ~/Desktop/MMCVRPTW-MLT
    source .venv/bin/activate
    python3 revise_master_round2.py
"""

import shutil
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook


MASTER_PATH = Path.home() / "Desktop" / "MMCVRPTW-MLT" / "MMCVRPTW_MLT_MasterData_V4.xlsx"
BACKUP_PATH = MASTER_PATH.with_suffix(".xlsx.bak2")


# ============================================================================
#                          ROUND 2 PRICING TABLES
# ============================================================================

# Lower FTL rates — was too aggressive in Round 1. Solver wouldn't use trucks.
# These are still realistic ₹/km × typical distance, but on the cheaper end
# of the published Indian freight indices for fleet contracts.
NEW_FTL_RATES = {
    "40ft_trailer":   38000,   # Round 1: 52000  (₹35/km × 1100km)
    "20ft_container": 22000,   # Round 1: 30000  (₹30/km × 750km)
    "14ft_van":        6000,   # Round 1:  7500  (₹22/km × 270km)
    "Tata_Ace":        2600,   # Round 1:  3200  (₹16/km × 160km)
}

NEW_PTL_RATES = {
    "40ft_trailer":   25000,   # ~65% of new FTL
    "20ft_container": 14500,
    "14ft_van":        4000,
    "Tata_Ace":        1700,
}

# LTL down a bit — Round 1 was ₹28; bring back closer to Round 0 (₹22)
NEW_LTL_PER_KG = {
    "14ft_van":  22,
    "Tata_Ace":  18,
}

# Penalty config — moderate-high to make breach genuinely undesirable
# but not infinite (so impossible orders are still routed at a cost).
NEW_PENALTY_CONFIG = {
    # (SLA_Tier, Priority): (WISMO, Compensation, Repurchase_Risk, Total)
    ("Same-day", "Prime"):    (2000, 5000, 8000, 15000),  # was 680
    ("Same-day", "Standard"): (1000, 3000, 4000,  8000),  # was 290
    ("Next-day", "Prime"):    (1500, 3500, 5000, 10000),  # was 440
    ("Next-day", "Standard"): ( 800, 2000, 3000,  5800),  # was 180
    ("2-day",    "Prime"):    ( 600, 1500, 2000,  4100),  # was 140
    ("2-day",    "Standard"): ( 400, 1000, 1500,  2900),  # was  80
}

# SLA_Config rewrite — distance-aware deadlines that match Indian ground reality.
# Same-City: < 50km, light traffic, 10/14h is achievable
# Same-Region: 200-500km, mostly highway, 18/24h is achievable
# Metro-to-Metro: 500-1500km, mixed routes, RAISE TO 48h Prime / 72h Standard
# Long-haul: 1500km+, two-day truck haul, NEW BAND 72h Prime / 96h Standard
NEW_SLA_BANDS = {
    # band_name: (Prime_hrs, Standard_hrs)
    "Same-City":      (10, 14),       # unchanged
    "Same-Region":    (24, 36),       # was 18/24 — bump for safety
    "Metro-to-Metro": (48, 72),       # was 28/36 — was too tight
    "Long-haul":      (72, 96),       # new band for >1500km pairs
}

# Pairs to RECLASSIFY as Long-haul (currently labeled Metro-to-Metro but >1500km).
# Determined by inspecting the lane matrix max-transit-hr; transit_hr > 40h
# corresponds to road distances >1500km at ~40 km/h avg highway speed.
# Format: set of (FC_Region, DS_City) tuples
LONG_HAUL_PAIRS = {
    # SOUTH → north/east destinations
    ("South", "GUW"), ("South", "CHA"), ("South", "DEL"), ("South", "BHU"),
    ("South", "LKN"), ("South", "KOL"), ("South", "JPR"), ("South", "AHM"),
    # EAST → south/west destinations
    ("East", "COK"), ("East", "MAD"), ("East", "COI"), ("East", "BLR"),
    ("East", "MUM"), ("East", "PUN"), ("East", "AHM"), ("East", "CHE"),
    ("East", "HYD"), ("East", "IND"), ("East", "JPR"), ("East", "CHA"),
    # WEST → far east/north-east
    ("West", "GUW"), ("West", "KOL"), ("West", "MAD"),
    # NORTH → south
    ("North", "COK"), ("North", "MAD"), ("North", "COI"),
    ("North", "CHE"), ("North", "BLR"), ("North", "GUW"),
}


# ============================================================================
#                           REVISION LOGIC
# ============================================================================

def revise_carrier_master(ws) -> int:
    changed = 0
    header_row = None
    for r in range(1, 5):
        if ws.cell(r, 1).value == "Carrier_ID":
            header_row = r
            break
    headers = {ws.cell(header_row, c).value: c for c in range(1, ws.max_column + 1)
               if ws.cell(header_row, c).value}
    col_v, col_l = headers["Vehicle_Type"], headers["Load_Type"]
    col_ftl = headers["FTL_Rate_INR_per_trip"]
    col_ptl = headers["PTL_Rate_INR_per_trip"]
    col_ltl = headers["LTL_Rate_INR_per_kg"]
    for r in range(header_row + 1, ws.max_row + 1):
        vehicle, loadtype = ws.cell(r, col_v).value, ws.cell(r, col_l).value
        if not vehicle or not loadtype:
            continue
        old, new, col = None, None, None
        if loadtype == "FTL" and vehicle in NEW_FTL_RATES:
            old, new, col = ws.cell(r, col_ftl).value, NEW_FTL_RATES[vehicle], col_ftl
        elif loadtype == "PTL" and vehicle in NEW_PTL_RATES:
            old, new, col = ws.cell(r, col_ptl).value, NEW_PTL_RATES[vehicle], col_ptl
        elif loadtype == "LTL" and vehicle in NEW_LTL_PER_KG:
            old, new, col = ws.cell(r, col_ltl).value, NEW_LTL_PER_KG[vehicle], col_ltl
        if col is not None and old != new:
            ws.cell(r, col).value = new
            print(f"  Row {r}: {vehicle}/{loadtype} — {old} -> {new}")
            changed += 1
    return changed


def revise_penalty_config(ws) -> int:
    changed = 0
    header_row = None
    for r in range(1, 5):
        if ws.cell(r, 1).value == "SLA_Tier":
            header_row = r
            break
    headers = {ws.cell(header_row, c).value: c for c in range(1, ws.max_column + 1)
               if ws.cell(header_row, c).value}
    cols = (headers["WISMO_Cost_INR"], headers["Compensation_INR"],
            headers["Repurchase_Risk_INR"], headers["Total_Penalty_INR"])
    for r in range(header_row + 1, ws.max_row + 1):
        tier, priority = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not tier or not priority:
            continue
        key = (tier, priority)
        if key not in NEW_PENALTY_CONFIG:
            continue
        wismo, comp, repurch, total = NEW_PENALTY_CONFIG[key]
        old_total = ws.cell(r, cols[3]).value
        ws.cell(r, cols[0]).value = wismo
        ws.cell(r, cols[1]).value = comp
        ws.cell(r, cols[2]).value = repurch
        ws.cell(r, cols[3]).value = total
        if old_total != total:
            print(f"  Row {r}: {tier}/{priority} — total {old_total} -> {total}")
            changed += 1
    return changed


def revise_sla_config(ws) -> int:
    """Update SLA_Config: reclassify long-haul pairs and set new band hours."""
    changed = 0
    header_row = None
    for r in range(1, 5):
        if ws.cell(r, 1).value == "FC_Region":
            header_row = r
            break
    for r in range(header_row + 1, ws.max_row + 1):
        reg, city = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not reg or not city:
            continue
        band = ws.cell(r, 3).value
        # If pair is in long-haul set, reclassify
        if (reg, city) in LONG_HAUL_PAIRS:
            new_band = "Long-haul"
        else:
            new_band = band  # keep current band
        pri_h, std_h = NEW_SLA_BANDS[new_band]
        old_band = ws.cell(r, 3).value
        old_pri = ws.cell(r, 4).value
        old_std = ws.cell(r, 5).value
        ws.cell(r, 3).value = new_band
        ws.cell(r, 4).value = pri_h
        ws.cell(r, 5).value = std_h
        if (old_band, old_pri, old_std) != (new_band, pri_h, std_h):
            print(f"  Row {r}: {reg}/{city} — {old_band} {old_pri}/{old_std}h -> {new_band} {pri_h}/{std_h}h")
            changed += 1
    return changed


def recompute_order_sla(wb) -> int:
    """Re-derive SLA_hrs, SLA_Tier, Customer_SLA_Deadline, Ops_SLA_Deadline,
    and penalty columns for every order based on the new SLA_Config.
    Returns # orders changed."""
    # Build lookup from updated SLA_Config: (region, city) -> (band, pri_h, std_h)
    ws_sla = wb["SLA_Config"]
    sla_lookup = {}
    header_row = None
    for r in range(1, 5):
        if ws_sla.cell(r, 1).value == "FC_Region":
            header_row = r
            break
    for r in range(header_row + 1, ws_sla.max_row + 1):
        reg, city = ws_sla.cell(r, 1).value, ws_sla.cell(r, 2).value
        if reg and city:
            band = ws_sla.cell(r, 3).value
            pri_h = ws_sla.cell(r, 4).value
            std_h = ws_sla.cell(r, 5).value
            sla_lookup[(reg, city)] = (band, pri_h, std_h)

    # Penalty lookup: (band-tier, priority) -> (wismo, comp, repurch, total)
    # Band -> SLA_Tier mapping
    band_to_tier = {
        "Same-City":      "Same-day",
        "Same-Region":    "Next-day",
        "Metro-to-Metro": "Next-day",
        "Long-haul":      "2-day",
    }

    ws_o = wb["Order_Data"]
    # Find columns
    col_placed = 7         # Order_Placed_Time
    col_cust_dl = 8        # Customer_SLA_Deadline
    col_priority = 9       # Priority
    col_region = 12        # FC_Region
    col_city = 13          # DS_City
    col_sla_hrs = 14       # SLA_hrs
    col_sla_tier = 15      # SLA_Tier
    col_wismo = 16
    col_comp = 17
    col_repurch = 18
    col_total = 19
    col_ops_dl = 20        # Ops_SLA_Deadline (= cust_dl - some buffer; we'll keep
                           #                   the existing buffer pattern)

    changed = 0
    for r in range(3, ws_o.max_row + 1):
        oid = ws_o.cell(r, 1).value
        if not oid:
            continue
        placed = ws_o.cell(r, col_placed).value
        priority = ws_o.cell(r, col_priority).value
        region = ws_o.cell(r, col_region).value
        city = ws_o.cell(r, col_city).value
        if not (placed and priority and region and city):
            continue

        # Look up SLA band & hours
        if (region, city) not in sla_lookup:
            continue
        band, pri_h, std_h = sla_lookup[(region, city)]
        new_sla_hrs = pri_h if priority == "Prime" else std_h
        new_tier = band_to_tier[band]

        # Recompute Customer_SLA_Deadline = order_placed + sla_hrs
        if isinstance(placed, datetime):
            new_cust_dl = placed + timedelta(hours=new_sla_hrs)
        else:
            new_cust_dl = ws_o.cell(r, col_cust_dl).value  # leave alone if not parseable

        # Ops_SLA_Deadline: use same offset pattern as before (typically 2h buffer)
        old_cust_dl = ws_o.cell(r, col_cust_dl).value
        old_ops_dl = ws_o.cell(r, col_ops_dl).value
        if isinstance(old_cust_dl, datetime) and isinstance(old_ops_dl, datetime):
            buffer = old_cust_dl - old_ops_dl  # preserve original buffer
        else:
            buffer = timedelta(hours=0)
        new_ops_dl = new_cust_dl - buffer if isinstance(new_cust_dl, datetime) else old_ops_dl

        # Look up new penalty values
        pen_key = (new_tier, priority)
        if pen_key in NEW_PENALTY_CONFIG:
            wismo, comp, repurch, total = NEW_PENALTY_CONFIG[pen_key]
        else:
            wismo = comp = repurch = total = None

        # Write
        ws_o.cell(r, col_sla_hrs).value = new_sla_hrs
        ws_o.cell(r, col_sla_tier).value = new_tier
        ws_o.cell(r, col_cust_dl).value = new_cust_dl
        if isinstance(new_ops_dl, datetime):
            ws_o.cell(r, col_ops_dl).value = new_ops_dl
        if wismo is not None:
            ws_o.cell(r, col_wismo).value = wismo
            ws_o.cell(r, col_comp).value = comp
            ws_o.cell(r, col_repurch).value = repurch
            ws_o.cell(r, col_total).value = total

        changed += 1

    return changed


def main():
    if not MASTER_PATH.exists():
        raise FileNotFoundError(f"Master not found at {MASTER_PATH}")

    print(f"Backing up master to {BACKUP_PATH}")
    shutil.copy2(MASTER_PATH, BACKUP_PATH)

    print(f"\nLoading {MASTER_PATH}")
    wb = load_workbook(MASTER_PATH)

    print(f"\n--- Revising Carrier_Master (FTL/PTL/LTL rates DOWN ~25%) ---")
    cm_changes = revise_carrier_master(wb["Carrier_Master"])

    print(f"\n--- Revising Penalty_Config (UP to ₹2.9k-15k for strong SLA preference) ---")
    pc_changes = revise_penalty_config(wb["Penalty_Config"])

    print(f"\n--- Revising SLA_Config (Metro-to-Metro 28/36 -> 48/72; new Long-haul band 72/96) ---")
    sla_changes = revise_sla_config(wb["SLA_Config"])

    print(f"\n--- Recomputing Order_Data SLA columns to match new bands ---")
    order_changes = recompute_order_sla(wb)

    print(f"\nSaving back to {MASTER_PATH}")
    wb.save(MASTER_PATH)

    print(f"\n=== DONE ===")
    print(f"  Carrier_Master cells changed: {cm_changes}")
    print(f"  Penalty_Config rows changed:  {pc_changes}")
    print(f"  SLA_Config rows changed:      {sla_changes}")
    print(f"  Order_Data orders updated:    {order_changes}")
    print(f"  Backup: {BACKUP_PATH}")
    print()
    print("Next:")
    print("  1. pytest backend/tests/ -v --tb=short  (should still pass 6/6)")
    print("  2. ./run.sh — run Balanced solve")
    print("  3. Send the output Excel")


if __name__ == "__main__":
    main()
